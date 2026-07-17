"""Build the GST Utilization Entry worksheet from a GSTR 3B workbook."""

from __future__ import annotations

import re
from numbers import Number

from openpyxl.styles import Alignment, Font


MONTHS_LONG = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]
MONTHS_SHORT = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]

ROW_LABELS = {
    "OUTWARD_TAXABLE": "(a) outward taxable supplies (other than zero rated, nilrated and exempted)",
    "INWARD_RCM": "(d) inward supplies (liable to reverse charge)",
    "IGST_FOR_IGST": "igst for igst",
    "CGST_FOR_CGST": "cgst for cgst",
    "SGST_FOR_SGST": "sgst for sgst",
    "CGST_DISCHARGE_IGST": "cgst utilised to discharge igst",
    "SGST_DISCHARGE_IGST": "sgst utilised to discharge",
    "IGST_DISCHARGE_CGST": "igst utilised to discharge cgst",
    "IGST_DISCHARGE_SGST": "igst utilised to discharge sgst",
    "CASH_UTILISED": "cash utilised cash balance",
}

# key, label, account code, side, manual
ACCOUNT_LINES = [
    ("IGST_OUT", "IGST Output", "1242010", "Debit", False),
    ("SGST_OUT", "SGST Output", "1242020", "Debit", False),
    ("CGST_OUT", "CGST Output", "1242030", "Debit", False),
    ("CESS_OUT", "Cess Output", "1242100", "Debit", False),
    ("IGST_RCM", "IGST Output - RCM", "1242040", "Debit", False),
    ("SGST_RCM", "SGST Output - RCM", "1242050", "Debit", False),
    ("CGST_RCM", "CGST Output - RCM", "1242060", "Debit", False),
    ("IGST_IN", "IGST Input", "2515010", "Credit", False),
    ("SGST_IN", "SGST Input", "2515030", "Credit", False),
    ("CGST_IN", "CGST Input", "2515020", "Credit", False),
    ("ICICI", "ICICI Bank", "2332420", "Credit", True),
    ("TCS_IGST", "TCS Collected - IGST", "2515090", "Credit", False),
    ("TCS_SGST", "TCS Collected - SGST", "2515080", "Credit", False),
    ("TCS_CGST", "TCS Collected - CGST", "2515070", "Credit", False),
    ("INTEREST", "Interest on Late payment of Taxes", "7231010", "", True),
    ("WRITTEN_OFF", "Written off", "7286020", "", True),
]

GSTIN_RE = re.compile(r"^\d{2}[A-Z0-9]{10,13}$", re.IGNORECASE)


def _find_source(workbook):
    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith("GSTR 3B Computation"):
            return workbook[sheet_name]
    raise ValueError("Could not find a sheet starting with 'GSTR 3B Computation'")


def _find_rows(source) -> dict[str, int]:
    rows: dict[str, int] = {}
    labels = {
        row: str(source.cell(row=row, column=1).value or "").strip().casefold()
        for row in range(1, 91)
    }
    for key, wanted in ROW_LABELS.items():
        wanted = wanted.casefold()
        match = next((row for row, label in labels.items() if wanted in label), None)
        if match is None:
            raise ValueError(f"Could not find row for: {key}")
        rows[key] = match
    return rows


def _find_blocks(source) -> list[dict]:
    blocks = []
    for col in range(1, source.max_column + 1):
        value = source.cell(row=6, column=col).value
        gstin = str(value or "").strip().upper()
        if GSTIN_RE.fullmatch(gstin):
            blocks.append({
                "state": str(source.cell(row=5, column=col).value or "").strip(),
                "gstin": gstin,
                "col": col,
            })
    if not blocks:
        raise ValueError("No GSTIN blocks found in the source sheet")
    return blocks


def _value(source, row: int, col: int, offset: int):
    value = source.cell(row=row, column=col + offset).value
    return value if isinstance(value, Number) and not isinstance(value, bool) else 0


def _narration_month(year_value, month_value) -> str:
    try:
        year = int(year_value)
        filing_index = MONTHS_LONG.index(str(month_value).strip().title())
    except (TypeError, ValueError) as exc:
        raise ValueError("Year in C1 and month name in C2 are required") from exc
    period_index = (filing_index - 1) % 12
    period_year = year if filing_index > 0 else year - 1
    return f"{MONTHS_SHORT[period_index]}-{period_year % 100:02d}"


def run(workbook, values_workbook=None) -> dict:
    """Modify *workbook* in place and return metadata needed by the scraper.

    ``values_workbook`` is the same uploaded file opened with ``data_only``.
    The Consultant sheet uses Excel formulas for ITC and cash utilisation;
    reading from that workbook gives their saved calculated values while the
    normal workbook remains intact for the completed download.
    """
    source = _find_source(values_workbook or workbook)
    narration_month = _narration_month(source.cell(1, 3).value, source.cell(2, 3).value)
    rows = _find_rows(source)
    blocks = _find_blocks(source)

    if "GST Utilization Entry" in workbook.sheetnames:
        del workbook["GST Utilization Entry"]
    target = workbook.create_sheet("GST Utilization Entry")
    # Mirror the existing Excel entry template.  It deliberately keeps the
    # standard gridline/Calibri look rather than adding colours or borders.
    bold = Font(bold=True)
    gstin_font = Font(color="0563C1")
    new_font = Font(color="FF0000")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    amount_format = '#,##0.00;-#,##0.00;-'
    target.sheet_view.showGridLines = True
    target.cell(3, 1, "GST Number").alignment = left
    target.cell(3, 4, "New").font = new_font
    target.cell(3, 4).alignment = left
    # These are the column positions and proportions of the existing entry
    # sheet.  B and G are intentionally blank working/separator columns.
    for column, width in {
        "A": 22, "B": 11, "C": 31, "D": 46, "E": 15,
        "F": 15, "G": 15, "H": 56, "I": 8, "J": 8,
    }.items():
        target.column_dimensions[column].width = width

    current_row = 4
    icici_rows: dict[str, int] = {}
    for block in blocks:
        col = block["col"]
        amounts = {
            "IGST_OUT": _value(source, rows["OUTWARD_TAXABLE"], col, 1),
            "SGST_OUT": _value(source, rows["OUTWARD_TAXABLE"], col, 3),
            "CGST_OUT": _value(source, rows["OUTWARD_TAXABLE"], col, 2),
            "CESS_OUT": _value(source, rows["OUTWARD_TAXABLE"], col, 4),
            "IGST_RCM": _value(source, rows["INWARD_RCM"], col, 1),
            "SGST_RCM": _value(source, rows["INWARD_RCM"], col, 3),
            "CGST_RCM": _value(source, rows["INWARD_RCM"], col, 2),
            "IGST_IN": round(-(
                _value(source, rows["IGST_FOR_IGST"], col, 1)
                + _value(source, rows["IGST_DISCHARGE_CGST"], col, 2)
                + _value(source, rows["IGST_DISCHARGE_SGST"], col, 3)
            ), 2),
            "CGST_IN": round(-(
                _value(source, rows["CGST_FOR_CGST"], col, 2)
                + _value(source, rows["CGST_DISCHARGE_IGST"], col, 1)
            ), 2),
            "SGST_IN": round(-(
                _value(source, rows["SGST_FOR_SGST"], col, 3)
                + _value(source, rows["SGST_DISCHARGE_IGST"], col, 1)
            ), 2),
            "TCS_IGST": _value(source, rows["CASH_UTILISED"], col, 1),
            "TCS_SGST": _value(source, rows["CASH_UTILISED"], col, 3),
            "TCS_CGST": _value(source, rows["CASH_UTILISED"], col, 2),
        }

        header_cells = ((1, block["gstin"], left), (3, block["state"], left), (5, "Debit", left), (6, "Credit", left))
        for cell, value, alignment in header_cells:
            target.cell(current_row, cell, value).font = gstin_font if cell == 1 else Font()
            target.cell(current_row, cell).alignment = alignment
        first_account_row = current_row + 1
        for index, (key, label, account_code, side, manual) in enumerate(ACCOUNT_LINES):
            row = first_account_row + index
            target.cell(row, 3, label)
            target.cell(row, 3).alignment = left
            target.cell(row, 4, f"101.999999999.{account_code}.999.{block['gstin'][:2]}.999.999.99999.9999999")
            target.cell(row, 4).alignment = left
            target.cell(row, 8, f"GST-{label.strip()}- Payment {narration_month}")
            target.cell(row, 8).alignment = left
            for amount_col in (5, 6):
                target.cell(row, amount_col).alignment = right
                target.cell(row, amount_col).number_format = amount_format
            amount = amounts.get(key, 0)
            if not manual:
                target.cell(row, 5 if side == "Debit" else 6, amount)
            if key == "ICICI":
                icici_rows[block["gstin"]] = row
        totals_row = first_account_row + 16
        for amount_col, formula in (
            (5, f"=SUM(E{first_account_row}:E{first_account_row + 15})"),
            (6, f"=SUM(F{first_account_row}:F{first_account_row + 15})"),
        ):
            target.cell(totals_row, amount_col, formula).font = bold
            target.cell(totals_row, amount_col).alignment = right
            target.cell(totals_row, amount_col).number_format = amount_format
        # The reference has two ordinary blank rows between state blocks.
        current_row += 20

    # Excel displays the filter drop-downs in the first state header row while
    # retaining every subsequent state block beneath the same filter range.
    target.auto_filter.ref = f"A4:H{current_row - 1}"

    return {
        "narration_month": narration_month,
        "states_found": len(blocks),
        "icici_rows": icici_rows,
    }
