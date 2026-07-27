"""Local GST receipt-to-workbook automation application."""

from __future__ import annotations

import io
import secrets
import time
from pathlib import Path
from urllib.parse import quote

import openpyxl
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

import computation
import receipts


ROOT = Path(__file__).parent
app = FastAPI(title="GST Challan Receipt Automation")
app.mount("/static", StaticFiles(directory=ROOT / "static"), name="static")
templates = Jinja2Templates(directory=ROOT / "templates")

DOWNLOAD_TTL_SECONDS = 30 * 60
completed_workbooks: dict[str, dict] = {}


def _discard_expired_workbooks() -> None:
    """Remove completed uploads that were not downloaded within 30 minutes."""
    now = time.monotonic()
    for token, item in list(completed_workbooks.items()):
        if item["expires_at"] <= now:
            del completed_workbooks[token]


def _store_completed_workbook(workbook, filename: str) -> str:
    """Return an unguessable, short-lived URL for one user's workbook."""
    _discard_expired_workbooks()
    token = secrets.token_urlsafe(32)
    completed_workbooks[token] = {
        "workbook": workbook,
        "filename": filename,
        "expires_at": time.monotonic() + DOWNLOAD_TTL_SECONDS,
    }
    return f"/download/{token}"


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.post("/upload")
async def upload(
    file: UploadFile = File(...),
    receipts_files: list[UploadFile] = File(..., alias="receipts"),
):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        return JSONResponse({"ok": False, "error": "Please upload an Excel (.xlsx) file"}, status_code=400)
    if not receipts_files:
        return JSONResponse({"ok": False, "error": "Please upload at least one GST challan receipt PDF"}, status_code=400)

    try:
        workbook_bytes = await file.read()
        # Keep one copy with formulas/formatting for the download and read a
        # second copy using Excel's saved formula results for the Consultant
        # sheet values that feed the Credit entries.
        workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes))
        values_workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)
        result = computation.run(workbook, values_workbook)
    except ValueError as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)
    except Exception as exc:
        return JSONResponse({"ok": False, "error": f"Could not read Excel file: {exc}"}, status_code=400)

    receipt_results: list[dict] = []
    seen_gstins: set[str] = set()
    written = 0
    for receipt_file in receipts_files:
        filename = Path(receipt_file.filename or "receipt.pdf").name
        if not filename.lower().endswith(".pdf"):
            receipt_results.append({"filename": filename, "status": "error", "message": "Not a PDF file"})
            continue
        try:
            receipt = receipts.extract_receipt(await receipt_file.read(), filename)
        except ValueError as exc:
            receipt_results.append({"filename": filename, "status": "error", "message": str(exc)})
            continue

        gstin = receipt["gstin"]
        if gstin in seen_gstins:
            receipt_results.append({**receipt, "status": "duplicate", "message": "Duplicate receipt for this GSTIN - not written"})
            continue
        seen_gstins.add(gstin)
        row = result["icici_rows"].get(gstin)
        if row is None:
            receipt_results.append({**receipt, "status": "unmatched", "message": "GSTIN is not present in the uploaded workbook"})
            continue

        workbook["GST Utilization Entry"].cell(row=row, column=6).value = receipt["amount"]
        receipt_results.append({**receipt, "status": "written", "message": "Amount written to ICICI Bank credit"})
        written += 1

    filename = Path(file.filename).name
    download_url = _store_completed_workbook(workbook, filename)
    return {
        "ok": True,
        "filename": filename,
        "narration_month": result["narration_month"],
        "states_found": result["states_found"],
        "receipts_uploaded": len(receipts_files),
        "written": written,
        "unmatched": sum(item["status"] == "unmatched" for item in receipt_results),
        "duplicates": sum(item["status"] == "duplicate" for item in receipt_results),
        "errors": sum(item["status"] == "error" for item in receipt_results),
        "results": receipt_results,
        "download_url": download_url,
    }


@app.get("/download/{token}")
async def download(token: str):
    _discard_expired_workbooks()
    item = completed_workbooks.get(token)
    if item is None:
        raise HTTPException(status_code=404, detail="This download link has expired - please upload again")
    buffer = io.BytesIO()
    item["workbook"].save(buffer)
    buffer.seek(0)
    safe_name = f"{Path(item['filename'] or 'gst_utilization').stem}_completed.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(safe_name)}"},
    )
